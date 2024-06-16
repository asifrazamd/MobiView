import re
import logging
from Lib.ParserUtils import ParserUtils
from SpecsInfo.MobileSpecsInfo import MobileSpecsInfo
from openpyxl.utils import get_column_letter
from Lib.FileSystemUtils import FileSystemUtils

log = logging.getLogger(__name__)


class CameraSpecsInfo(MobileSpecsInfo):

    def _init_(self):

        super()._init_()
        self.CameraResolution = None
        self.CameraShutter = None

    def grepInfo(self):
            self.CameraResolution()
            self.CameraShutter()

            return self.CameraSpecsInfoDict

    def cleanup(self):
        pass

    def CameraResolution(self):
        ''' @function: getDeviceSerialNo gets Device Serial Number
        @param: None
        @return: Serial Number of Device '''
        self.command = self.ADBObj.getADBDumpsysCommand() + ' media.camera | grep  "Resolution"'
        Resolution = self.executeCommandOnDevice(command=self.command)
        # resolution = output.split(": ")[1].strip()
        print("camera")

        print(Resolution)
        self.updateDictionary(dictName=self.CameraSpecsInfoDict, key='Resolution', value=Resolution)
        return Resolution
    def CameraShutter(self):
        ''' @function: getDeviceSerialNo gets Device Serial Number
        @param: None
        @return: Serial Number of Device '''
        self.command = self.ADBObj.getADBDumpsysCommand() + ' media.camera | grep /I  "Shutter Range"'
        Shutter = self.executeCommandOnDevice(command=self.command)
        # resolution = output.split(": ")[1].strip()
        print("camera")

        print(Shutter)
        self.updateDictionary(dictName=self.CameraSpecsInfoDict, key='Shutter Range', value=Shutter)
        return Shutter

    def generateXLSXReport(self, xlsObj=None, wb=None, ws=None, dataDict=None):
        headers = []
        headers.insert(0, "Parameters")
        headers.insert(1, "Results")

        for idx in range(0, len(headers)):
            cellref = ws.cell(row=2, column=idx + 2)
            ws.column_dimensions[get_column_letter(idx + 2)].width = 40
            cellref.style = xlsObj.getNamedStyle(stylename="headerRow")

            cellref.value = headers[idx]

        dictkeys = list(dataDict.keys())
        for idx in range(0, len(dictkeys)):
            cellref = ws.cell(row=idx + 3, column=2)
            cellref.style = xlsObj.getNamedStyle(stylename="normalRow")
            cellref.value = dictkeys[idx]

        # headers = list(dataDict.keys())
        col_idx = 3
        row_idx = 3
        for datavalue in dataDict.values():
            cellref = ws.cell(row=row_idx, column=col_idx)
            cellref.style = xlsObj.getNamedStyle(stylename="normalRow")
            charlist = ["[", "'", "]"]
            datavalue = FileSystemUtils.replaceChars(datavalue, charlist)
            cellref.value = str(datavalue)
            row_idx += 1

        col_idx = 2
        row_idx = len(list(dataDict.keys())) + 2
        for ctr in range(col_idx, col_idx + 2):
            cellref = ws.cell(row=row_idx, column=ctr)
            cellref.style = xlsObj.getNamedStyle(stylename="lastRow")