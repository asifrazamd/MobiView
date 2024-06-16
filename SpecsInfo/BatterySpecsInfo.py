# Getting Battery Specifications

# Importing necessary libraries and modules

import re
import logging
from Lib.ParserUtils import ParserUtils
from SpecsInfo.MobileSpecsInfo import MobileSpecsInfo
from openpyxl.utils import get_column_letter
from Lib.FileSystemUtils import FileSystemUtils

# Creating a logger

log = logging.getLogger(__name__)


class BatterySpecsInfo(MobileSpecsInfo):

    def __init__(self):

        super().__init__()

        # Initialize battery specifications attributes to None

        self.BatteryVoltage = None      # Voltage of the battery
        self.BatteryStatus = None       # Status of the battery (e.g. charging, discharging, etc.)
        self.BatteryHealth = None       # Health of the battery (e.g. good, bad, etc.)
        self.BatteryLevel = None        # Current level of the battery (e.g. 50%, 75%, etc.)
        self.BatteryScale = None        # Scale of the battery (e.g. 100mAh, 5000mAh, etc.)
        self.BatteryTemperature = None  # Temperature of the battery
        self.BatteryTechnology = None   # Technology of the battery (e.g. Lithium-ion, Nickel-cadmium, etc.)



    def grepInfo(self):

        self.getBatteryVoltage()        # Get battery voltage and store in BatterySpecsInfoDict
        self.getBatteryStatus()         # Get battery status and store in BatterySpecsInfoDict
        self.getBatteryHealth()         # Get battery health and store in BatterySpecsInfoDict
        self.getBatteryLevel()          # Get battery level and store in BatterySpecsInfoDict
        self.getBatteryScale()          # Get battery scale and store in BatterySpecsInfoDict
        self.getBatteryTemperature()    # Get battery temperature and store in BatterySpecsInfoDict
        self.getBatteryTechnology()     # Get battery technology and store in BatterySpecsInfoDict

        # Return the BatterySpecsInfoDict containing all the battery specs

        return self.BatterySpecsInfoDict

    def cleanup(self):
        pass

    def getBatteryVoltage(self):
        # Execute ADB command to get battery voltage

        output = self.executeCommandOnDevice(command=self.ADBObj.getBattery()+' "voltage" ')

        # Parse the output using regular expressions to extract the voltage value

        rvalue = ParserUtils.parseDataViaRegex(r'(?s).*voltage: \d+.*voltage: (?P<voltage>\d+)',output)

        # Convert voltage to desired units

        voltage=int(rvalue.get('voltage'))/1000

        # # Update the BatterySpecsInfoDict dictionary with the key 'Battery_Voltage' and the value 'a'

        self.updateDictionary(dictName=self.BatterySpecsInfoDict, key='Battery_Voltage',value=voltage)

        # Return the battery voltage

        return voltage
    def getBatteryStatus(self):

        '''
        Battery status values:

                    1. Charging
                    2. Discharging
                    3. Not Charging
                    4. Full
                    5. Unknown
                    6. Idle
                    7. Initializing
                    8. Unknown
        '''

        # Execute a command on the device to get the battery status

        output = self.executeCommandOnDevice(command=self.ADBObj.getBattery()+' "status" ')

        # Parse the output using regular expressions to extract the status value

        rvalue = ParserUtils.parseDataViaRegex(r'status: (?P<status>\w+)',output)

        # Update the BatterySpecsInfoDict dictionary with the key 'Battery_Status' and the value 'status'

        self.updateDictionary(dictName=self.BatterySpecsInfoDict, key='Battery_Status',value=rvalue.get('status'))

        # Return the status value

        return rvalue.get('status')

    def getBatteryHealth(self):
        '''
        Battery health values:

                    1. Good
                    2. Overheat
                    3. Dead
                    4. Over Voltage
                    5. Unspecified Failure
                    6. Cold
                    7.Weak
                    8.Unknown
        '''

        # Execute a command on the device to get the battery health

        output = self.executeCommandOnDevice(command=self.ADBObj.getBattery()+' "health" ')

        # Parse the output using regular expressions to extract the health value

        rvalue = ParserUtils.parseDataViaRegex(r'health: (?P<health>\w+)',output)

        # Update the BatterySpecsInfoDict dictionary with the key 'Battery_Health' and the value 'health'

        self.updateDictionary(dictName=self.BatterySpecsInfoDict, key='Battery_Health',value=rvalue.get('health'))

        # Return the health value

        return rvalue.get('health')

    def getBatteryLevel(self):

        # Execute a command on the device to get the battery level

        output = self.executeCommandOnDevice(command=self.ADBObj.getBattery()+' "level" ')

        # Parse the output using regular expressions to extract the level value

        rvalue = ParserUtils.parseDataViaRegex(r'level: (?P<level>\w+)',output)

        percentage=(rvalue.get("level"))+"%"

        # Update the BatterySpecsInfoDict dictionary with the key 'Battery_Level' and the value 'level'

        self.updateDictionary(dictName=self.BatterySpecsInfoDict, key='Battery_Level',value=percentage)

        # Return the level level

        return percentage
    def getBatteryScale(self):

        # Execute a command on the device to get the battery scale

        output = self.executeCommandOnDevice(command=self.ADBObj.getBattery()+' "scale" ')

        # Parse the output using regular expressions to extract the scale value

        rvalue = ParserUtils.parseDataViaRegex(r'scale: (?P<scale>\w+)',output)

        # Update the BatterySpecsInfoDict dictionary with the key 'Battery_Scale' and the value 'scale'

        self.updateDictionary(dictName=self.BatterySpecsInfoDict, key='Battery_Scale',value=rvalue.get('scale'))

        # Return the level value

        return rvalue.get('scale')
    def getBatteryTemperature(self):

        # Execute a command on the device to get the battery temperature

        output = self.executeCommandOnDevice(command=self.ADBObj.getBattery()+' "temperature" ')

        # Parse the output using regular expressions to extract the temperature value

        rvalue = ParserUtils.parseDataViaRegex(r'temperature: (?P<temperature>\w+)',output)

        # Convert the extracted temperature value to an integer, divide by 10, and convert to string

        temperature=str(int((rvalue.get('temperature')))/10) +"°C"

        # Update the BatterySpecsInfoDict dictionary with the key 'Battery_Temperature' and the value 'a'

        self.updateDictionary(dictName=self.BatterySpecsInfoDict, key='Battery_Temperature',value=temperature)

        # Return the temperature value as a string

        return temperature
    def getBatteryTechnology(self):

        # Execute a command on the device to get the battery technology

        output = self.executeCommandOnDevice(command=self.ADBObj.getBattery()+' "technology" ')

        # Parse the output using regular expressions to extract the technology value

        rvalue = ParserUtils.parseDataViaRegex(r'technology: (?P<technology>\w+-\w+)',output)

        # Update the BatterySpecsInfoDict dictionary with the key 'Battery_Technology' and the value 'technology'

        self.updateDictionary(dictName=self.BatterySpecsInfoDict, key='Battery_Technology',value=rvalue.get('technology'))

        # Return the technology value
        return rvalue.get('technology')

    '''def generateXLSXReport(self, xlsObj=None, wb=None, ws=None, dataDict=None):

            headers = [] # Initialize headers list and insert column headers

            headers.insert(0, "Parameters") # Column header for parameter names

            headers.insert(1, "Results") # Column header for parameter values


            # Set column widths and styles for headers

            for idx in range(0, len(headers)):
                cellref = ws.cell(row=2, column=idx + 2)
                ws.column_dimensions[get_column_letter(idx + 2)].width = 40
                cellref.style = xlsObj.getNamedStyle(stylename="headerRow")

                cellref.value = headers[idx]

            # Get list of keys from data dictionary

            dictkeys = list(dataDict.keys())

            # Write parameter names and values to worksheet

            for idx in range(0, len(dictkeys)):
                cellref = ws.cell(row=idx + 3, column=2)
                cellref.style = xlsObj.getNamedStyle(stylename="normalRow")
                cellref.value = dictkeys[idx]

            # Write parameter values to worksheet

            col_idx = 3
            row_idx = 3
            for datavalue in dataDict.values():
                cellref = ws.cell(row=row_idx, column=col_idx)
                cellref.style = xlsObj.getNamedStyle(stylename="normalRow")
                charlist = ["[", "'", "]"] # Remove unwanted characters from data value
                datavalue = FileSystemUtils.replaceChars(datavalue, charlist)
                cellref.value = str(datavalue)
                row_idx += 1

            # Set style for last row

            col_idx = 2
            row_idx = len(list(dataDict.keys())) + 2
            for ctr in range(col_idx, col_idx + 2):
                cellref = ws.cell(row=row_idx, column=ctr)
                cellref.style = xlsObj.getNamedStyle(stylename="lastRow")'''

    def generateXLSXReport(self, xlsObj=None, wb=None, ws=None, dataDict=None):
        headers = []
        # Initialize headers list and insert column headers
        headers.insert(0, "Parameters")
        headers.insert(1, "Results")
        headers.insert(2, "Description")

        # Set column widths and styles for headers
        for idx in range(0, len(headers)):
            cellref = ws.cell(row=2, column=idx + 2)
            ws.column_dimensions[get_column_letter(idx + 2)].width = 40
            cellref.style = xlsObj.getNamedStyle(stylename="headerRow")
            cellref.value = headers[idx]

        # Define dictionaries that map parameter names and values to descriptions
        param_descriptions = {
            "Battery_Voltage": "The current voltage of the battery.",
            "Battery_Status": "",
            "Battery_Health": "",
            "Battery_Level": "The current level of the battery ",
            "Battery_Scale": "The current scale of the battery ",
            "Battery_Temperature": "The current temperature of the battery.",
            "Battery_Technology": "The battery technology."
        }

        status_descriptions = {
            "1": "The current status of the battery is Charging",
            "2": "The current status of the battery is Discharging",
            "3": "The current status of the battery is Full",
            "4": "The current status of the battery is Not Charging",
            "5": "The current status of the battery is Unknown"
        }

        health_descriptions = {
            "1": "The current health of the battery is Good",
            "2": "The current health of the battery is Overheat",
            "3": "The current health of the battery is Dead",
            "4": "The current health of the battery is Over Voltage"
        }


        # Get list of keys from data dictionary
        dictkeys = list(dataDict.keys())

        # Write parameter names and values to worksheet
        for idx in range(0, len(dictkeys)):
            cellref = ws.cell(row=idx + 3, column=2)
            cellref.style = xlsObj.getNamedStyle(stylename="normalRow")
            cellref.value = dictkeys[idx]

            cellref = ws.cell(row=idx + 3, column=3)
            cellref.style = xlsObj.getNamedStyle(stylename="normalRow")
            charlist = ["[", "'", "]"]
            datavalue = FileSystemUtils.replaceChars(dataDict.get(dictkeys[idx]), charlist)
            cellref.value = str(datavalue)

            # Write parameter descriptions to worksheet
            cellref = ws.cell(row=idx + 3, column=4)
            cellref.style = xlsObj.getNamedStyle(stylename="normalRow")
            if dictkeys[idx] == "Battery_Status":
                cellref.value = status_descriptions.get(str(dataDict.get(dictkeys[idx])), "")
            elif dictkeys[idx] == "Battery_Health":
                cellref.value = health_descriptions.get(str(dataDict.get(dictkeys[idx])), "")
            else:
                cellref.value = param_descriptions.get(dictkeys[idx], "")

        # Set style for last row
        col_idx = 2
        row_idx = len(list(dataDict.keys())) + 2
        for ctr in range(col_idx, col_idx + 3):
            cellref = ws.cell(row=row_idx, column=ctr)
            cellref.style = xlsObj.getNamedStyle(stylename="lastRow")





